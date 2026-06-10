"""Unit tests for Excel tools."""

import os
import ssl
import httpx
import tempfile
from pathlib import Path

import openpyxl
import pytest

os.environ["PYTHONHTTPSVERIFY"] = "0"
os.environ["CURL_CA_BUNDLE"] = ""
os.environ["REQUESTS_CA_BUNDLE"] = ""
ssl._create_default_https_context = ssl._create_unverified_context


ROOT = Path(__file__).resolve().parents[1]
import sys
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def create_test_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Test Instances"
    ws1.append(["Test Name", "Method", "Parameter"])
    ws1.append(["IDDQ_TEST", "IDDQ", "default"])

    ws2 = wb.create_sheet("Limits")
    ws2.append(["Test Name", "Hi Limit", "Lo Limit", "Units"])
    ws2.append(["IDDQ_TEST", 100, 0, "uA"])

    ws3 = wb.create_sheet("Pin Map")
    ws3.append(["Pin", "Channel"])
    ws3.append(["VCC", "CH1"])

    ws4 = wb.create_sheet("Levels")
    ws4.append(["Level Name", "VCC", "VDD"])
    ws4.append(["nom", 3.3, 1.8])

    wb.save(path)


def test_read_excel_sheets() -> None:
    from tools.excel_tools import read_excel_sheets

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = handle.name
    try:
        create_test_workbook(path)
        result = read_excel_sheets(path)
        assert "Test Instances" in result
        assert "Limits" in result
        assert result["Limits"][0]["Test Name"] == "IDDQ_TEST"
        assert result["Limits"][0]["Hi Limit"] == 100
    finally:
        os.unlink(path)


def test_write_excel_cell_roundtrip() -> None:
    from tools.excel_tools import write_excel_cell, read_excel_sheets

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = handle.name
    try:
        create_test_workbook(path)
        result = write_excel_cell(path, "Limits", 2, 2, 150, backup=False)
        assert result["status"] == "ok"
        assert result["old_value"] == 100
        assert result["new_value"] == 150

        data = read_excel_sheets(path)
        assert data["Limits"][0]["Hi Limit"] == 150
    finally:
        os.unlink(path)


def test_write_pin_map_blocked() -> None:
    from tools.excel_tools import write_excel_cell

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = handle.name
    try:
        create_test_workbook(path)
        with pytest.raises(ValueError, match="read-only"):
            write_excel_cell(path, "Pin Map", 2, 1, "CH99", backup=False)
    finally:
        os.unlink(path)


def test_write_voltage_limit_blocked() -> None:
    from tools.excel_tools import write_excel_cell

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = handle.name
    try:
        create_test_workbook(path)
        with pytest.raises(ValueError, match="Voltage"):
            write_excel_cell(path, "Levels", 2, 2, 99.9, backup=False)
    finally:
        os.unlink(path)
