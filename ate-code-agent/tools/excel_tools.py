"""
Excel tools for reading and writing IG-XL test program Excel workbooks.
Uses openpyxl and does not require Excel installation.
"""

import os
import ssl
import httpx
from pathlib import Path
from typing import Any

import openpyxl

from safety.rules import SAFETY_RULES

os.environ["PYTHONHTTPSVERIFY"] = "0"
os.environ["CURL_CA_BUNDLE"] = ""
os.environ["REQUESTS_CA_BUNDLE"] = ""
ssl._create_default_https_context = ssl._create_unverified_context


SHEETS_TO_READ = [
    "Test Instances",
    "Test Suites",
    "Limits",
    "Flow",
    "Levels",
    "Timing",
    "Pin Map",
    "Spec",
]


def read_excel_sheets(file_path: str) -> dict:
    """Read IG-XL sheets into {sheet_name: [row_dict, ...]} format."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"IG-XL program file not found: {file_path}")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in SHEETS_TO_READ:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append(dict(zip(headers, row)))
        result[sheet_name] = rows

    wb.close()
    return result


def write_excel_cell(
    file_path: str,
    sheet_name: str,
    row: int,
    col: int,
    value: Any,
    backup: bool = True,
) -> dict:
    """Write one IG-XL cell with safety validation and optional git backup."""
    if sheet_name in SAFETY_RULES["readonly_sheets"]:
        raise ValueError(
            f"SAFETY BLOCK: '{sheet_name}' is read-only. "
            "Editing this sheet risks hardware damage."
        )

    if sheet_name == "Levels" and isinstance(value, (int, float)):
        max_v = SAFETY_RULES["voltage_limits"]["VCC_max"]
        if value > max_v:
            raise ValueError(
                f"SAFETY BLOCK: Voltage {value}V exceeds maximum {max_v}V. "
                "This could destroy DUT devices."
            )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"IG-XL file not found: {file_path}")

    if backup and SAFETY_RULES["git_backup_enabled"]:
        from tools.git_tools import git_commit_backup

        git_commit_backup(file_path, f"Pre-edit backup: {sheet_name} row={row} col={col}")

    wb = openpyxl.load_workbook(str(path))
    ws = wb[sheet_name]
    old_value = ws.cell(row=row, column=col).value
    ws.cell(row=row, column=col, value=value)
    wb.save(str(path))
    wb.close()

    return {
        "sheet": sheet_name,
        "row": row,
        "col": col,
        "old_value": old_value,
        "new_value": value,
        "status": "ok",
    }
